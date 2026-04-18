import os
from django.core.management.base import BaseCommand
from django.conf import settings
import openpyxl


IMPORT_DIR = os.path.join(
    settings.BASE_DIR.parent,
    'Прил_образец_КОД_09.02.07-2-2026', 'БУ', 'Модуль 1',
    'Прил_2_ОЗ_КОД 09.02.07-2-2026-М1', 'import'
)

MEDIA_PRODUCTS = os.path.join(settings.MEDIA_ROOT, 'products')


class Command(BaseCommand):
    help = 'Импорт данных из xlsx файлов'

    def handle(self, *args, **options):
        self._import_delivery_points()
        self._import_products()
        self._import_users()
        self._import_orders()
        self.stdout.write(self.style.SUCCESS('Импорт завершён успешно'))

    def _import_delivery_points(self):
        from shop.models import DeliveryPoint
        path = os.path.join(IMPORT_DIR, 'Пункты выдачи_import.xlsx')
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        created = 0
        for row in ws.iter_rows(min_row=1, values_only=True):
            address = row[0]
            if address:
                _, c = DeliveryPoint.objects.get_or_create(address=str(address).strip())
                if c:
                    created += 1
        self.stdout.write(f'  Пункты выдачи: создано {created}')

    def _import_products(self):
        from shop.models import Product
        path = os.path.join(IMPORT_DIR, 'tovar.xlsx')
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        created = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            article, name, unit, price, bakery, cake_type, category, discount, stock, description, photo_name = row

            photo_rel = None
            if photo_name:
                photo_path = os.path.join(MEDIA_PRODUCTS, photo_name)
                if os.path.exists(photo_path):
                    photo_rel = f'products/{photo_name}'

            product, c = Product.objects.update_or_create(
                article=str(article).strip(),
                defaults={
                    'name': str(name).strip(),
                    'unit': str(unit).strip() if unit else 'шт.',
                    'price': float(price),
                    'bakery': str(bakery).strip(),
                    'cake_type': str(cake_type).strip(),
                    'category': str(category).strip(),
                    'discount': int(discount) if discount else 0,
                    'stock': int(stock) if stock else 0,
                    'description': str(description).strip() if description else '',
                    'photo': photo_rel or '',
                }
            )
            if c:
                created += 1
        self.stdout.write(f'  Товары: создано {created}')

    def _import_users(self):
        from shop.models import User
        path = os.path.join(IMPORT_DIR, 'user_import.xlsx')
        wb = openpyxl.load_workbook(path)
        ws = wb.active

        role_map = {
            'Администратор': User.ROLE_ADMIN,
            'Менеджер': User.ROLE_MANAGER,
            'Авторизированный клиент': User.ROLE_CLIENT,
        }

        created = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            role_ru, full_name, email, password = row
            role = role_map.get(str(role_ru).strip(), User.ROLE_CLIENT)

            if not User.objects.filter(email=str(email).strip()).exists():
                user = User.objects.create_user(
                    email=str(email).strip(),
                    password=str(password).strip(),
                    full_name=str(full_name).strip(),
                    role=role,
                    is_staff=(role == User.ROLE_ADMIN),
                    is_superuser=(role == User.ROLE_ADMIN),
                )
                created += 1
        self.stdout.write(f'  Пользователи: создано {created}')

    def _import_orders(self):
        from shop.models import Order, OrderItem, DeliveryPoint, User, Product
        path = os.path.join(IMPORT_DIR, 'Заказ_import.xlsx')
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        created = 0

        delivery_points = list(DeliveryPoint.objects.all().order_by('id'))

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            number, articles_str, order_date, delivery_date, point_idx, customer_name, receipt_code, status = row

            if Order.objects.filter(number=int(number)).exists():
                continue

            point_idx = int(point_idx) if point_idx else 1
            if point_idx <= len(delivery_points):
                delivery_point = delivery_points[point_idx - 1]
            else:
                delivery_point = delivery_points[0]

            customer = User.objects.filter(
                full_name=str(customer_name).strip(), role=User.ROLE_CLIENT
            ).first()

            if hasattr(order_date, 'date'):
                order_date = order_date.date()
            if hasattr(delivery_date, 'date'):
                delivery_date = delivery_date.date()

            order = Order.objects.create(
                number=int(number),
                customer=customer,
                customer_name=str(customer_name).strip(),
                order_date=order_date,
                delivery_date=delivery_date,
                delivery_point=delivery_point,
                receipt_code=int(receipt_code),
                status=str(status).strip(),
            )

            if articles_str:
                parts = [p.strip() for p in str(articles_str).split(',')]
                i = 0
                while i < len(parts):
                    article = parts[i].strip()
                    qty = 1
                    if i + 1 < len(parts):
                        try:
                            qty = int(parts[i + 1].strip())
                            i += 2
                        except ValueError:
                            i += 1
                    else:
                        i += 1

                    product = Product.objects.filter(article=article).first()
                    if product:
                        OrderItem.objects.create(order=order, product=product, quantity=qty)

            created += 1
        self.stdout.write(f'  Заказы: создано {created}')
